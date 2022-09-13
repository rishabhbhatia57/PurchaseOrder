import pandas as pd
import os
import xlsxwriter
import time
from openpyxl import load_workbook,Workbook
import openpyxl.utils.cell
import shutil
from datetime import datetime
import log
import os



logger = log.setup_custom_logger('root')

def generatingPackaingSlip(RootFolder,ReqSource,OrderDate,ClientCode,Formulasheet,TemplateFiles):
    try:
        #converting str to datetime
        OrderDate = datetime.strptime(OrderDate, '%Y-%m-%d')
        # extracting year from the order date
        year = OrderDate.strftime("%Y")
        # formatting order date {2022-00-00) format
        OrderDate = OrderDate.strftime('%Y-%m-%d')

        startedTemplating = time.time()
        sourcePivot = ReqSource
        source = TemplateFiles+"/PackingSlip-Template.xlsx"
        destination = TemplateFiles+"/TemplateFile.xlsx"
        

        # Making Copy of template file
        # shutil.copy(source, destination)
        # print("File copied successfully.")

        # Load work vook and sheets
        InputWorkbook = load_workbook(sourcePivot,data_only=True)
        # TemplateWorkbook = load_workbook(destination)

        InputSheet = InputWorkbook.active
        # TemplateSheet = TemplateWorkbook['ORDER']

        # Get rows and Column count
        df = pd.DataFrame(InputSheet, index=None)
        rows = len(df.axes[0])
        cols = len(df.axes[1])


        formulaWorksheet = load_workbook(Formulasheet+'/FormulaSheet.xlsx',data_only=True)
        formulaSheet = formulaWorksheet['FormulaSheet']
        DBFformula = formulaWorksheet['DBF']
        
        
        
        for column in range(2,cols-3):
            startedTemplatingFile = time.time()
            # Making Copy of template file
            shutil.copy(source, destination)
            logger.info("Template File copied successfully for generating packaging-slip")

            # Load work vook and sheets
            TemplateWorkbook = load_workbook(destination, data_only=True)
            TemplateSheet = TemplateWorkbook['ORDER']
            dbfsheet = TemplateWorkbook['DBF']

            
            TemplateSheet.cell(6,4).value = InputSheet.cell(7,2).value # Order Name
            TemplateSheet.cell(5,1).value = InputSheet.cell(7,2).value # Order Name
            # PO Number
            filename = InputSheet.cell(4,column).value
            TemplateSheet.cell(5,2).value = InputSheet.cell(4,column).value
            TemplateSheet.cell(6,2).value = InputSheet.cell(4,column).value
            TemplateSheet.cell(6,1).value = InputSheet.cell(4,column).value
            TemplateSheet.cell(6,3).value = InputSheet.cell(4,column).value
            # Receving Location
            TemplateSheet.cell(5,3).value = InputSheet.cell(5,column).value
            TemplateSheet.cell(4,2).value = InputSheet.cell(5,column).value

            TemplateSheet.cell(1,1).value = 'DATE'
            TemplateSheet.cell(1,2).value = InputSheet.cell(2,4).value # Date

            TemplateSheet.cell(1,3).value = 'SGST/IGST'
            TemplateSheet.cell(1,4).value = InputSheet.cell(6,column).value # IGST/SGST Type
            print(TemplateSheet.cell(1,4).value,InputSheet.cell(6,column).value)

            # Copy EAN to template sheet
            Trows = 8
            Tcols = 5
            dbfrows = 2
            dbfcols = 57
            for row in range(7,rows):
                # if InputSheet.cell(row,column).value != None or InputSheet.cell(row,column).value != "":
                if str(InputSheet.cell(row,column).value).isnumeric():
                    
                        
                    # Copy Qty to template sheet
                    TemplateSheet.cell(Trows,Tcols).value = InputSheet.cell(row,column).value 
                    TemplateSheet.cell(Trows,Tcols+1).value = InputSheet.cell(row,column).value
                    
                    # Copy EAN to template sheet
                    TemplateSheet.cell(Trows,Tcols-3).value = InputSheet.cell(row,1).value

                    # VLOOKUP
                    # StyleName
                    TemplateSheet.cell(Trows,Tcols-4).value = "="+formulaSheet.cell(3,2).value.replace("#VAL#",str(Trows))

                    # style
                    TemplateSheet.cell(Trows,Tcols-2).value =  "="+formulaSheet.cell(4,2).value.replace("#VAL#",str(Trows))

                    # SADM SKU
                    TemplateSheet.cell(Trows,Tcols-1).value = "="+formulaSheet.cell(5,2).value.replace("#VAL#",str(Trows))
                    
                    # Rate (in Rs.) Order file
                    TemplateSheet.cell(Trows,Tcols+3).value = "="+formulaSheet.cell(6,2).value.replace("#VAL#",str(Trows))

                    # Closing stk
                    TemplateSheet.cell(Trows,Tcols+5).value = "="+formulaSheet.cell(11,2).value.replace("#VAL#",str(Trows)) 

                    #Cls stk vs order
                    # TemplateSheet.cell(Trows,Tcols+6).value = TemplateSheet.cell(Trows,Tcols+5).value - TemplateSheet.cell(Trows,Tcols+2).value
                    TemplateSheet.cell(Trows,Tcols+6).value = "="+openpyxl.utils.cell.get_column_letter(Tcols+5)+str(Trows) +'-'+openpyxl.utils.cell.get_column_letter(Tcols+2)+str(Trows) 

                    # LOCATION2
                    TemplateSheet.cell(Trows,Tcols+7).value = "="+formulaSheet.cell(7,2).value.replace("#VAL#",str(Trows))

                    #BULK  / DTA  BULK  /  EOSS LOC
                    TemplateSheet.cell(Trows,Tcols+8).value = "="+formulaSheet.cell(8,2).value.replace("#VAL#",str(Trows))

                    #MRP
                    TemplateSheet.cell(Trows,Tcols+9).value = "="+formulaSheet.cell(9,2).value.replace("#VAL#",str(Trows))
                    
                    # Adding values to DBF
                    for i in range(1,dbfcols):
                        dbfsheet.cell(dbfrows, i).value = '='+DBFformula.cell(2,i+1).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows))
                    
                    Trows += 1
                    dbfrows += 1 

            
            TemplateWorkbook.save(RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/70-Packaging-Slip/"+"PackagingSlip_"+str(filename)+".xlsx")
            logger.info("Packaing slip generated for: "+str(filename)+ " file - {:.2f}".format(time.time() - startedTemplatingFile,2)+ " seconds.")
            
        logger.info("Total time taken for generation of packaging-slips:  {:.2f}".format(time.time() - startedTemplating,2)+ " seconds.")
        return 'Completed!'
    except Exception as e:
        logger.error("Error while generating packaging-slip file: "+str(e))
