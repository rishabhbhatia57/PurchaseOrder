import pandas as pd
import numpy as np
import xlsxwriter
from datetime import datetime
from openpyxl import load_workbook,Workbook
import openpyxl.utils.cell
from datetime import datetime
import log
import os

logger = log.setup_custom_logger('root')

def mergeToPivot(RootFolder,POSource,OrderDate,ClientCode,Formulasheet):
    try:
        #converting str to datetime
        OrderDate = datetime.strptime(OrderDate, '%Y-%m-%d')
        # extracting year from the order date
        year = OrderDate.strftime("%Y")
        # formatting order date {2022-00-00) format
        OrderDate = OrderDate.strftime('%Y-%m-%d')

        formulaWorksheet = load_workbook(Formulasheet+'/FormulaSheet.xlsx',data_only=True) 
        # Data_only = True is used to get evaluated formula value instead of formula
        formulaSheet = formulaWorksheet['FormulaSheet']
        if not os.path.exists(RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/50-Consolidate-Orders/Consolidate-Orders.xlsx"):
            logger.info("Could not find the consolidate order folder to generate requirement summary file")
            return
        else:
            df = pd.read_excel(RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/50-Consolidate-Orders/Consolidate-Orders.xlsx")
            df_pivot = pd.pivot_table(df, index="ArticleEAN", values='Qty', 
            columns=['Vendor Name','PO Number','Receiving Location'], aggfunc='sum')
            df_pivot['Grand Total'] = 0
            df_pivot['Closing Stock'] = 0
            df_pivot['Diff CS - GT'] = 0
            df_pivot['Rate'] = 0
            df_pivot.to_excel(RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/60-Requirement-Summary/Requirement-Summary.xlsx")


            # Adding Processing Date, Order Number and Closing Stock, Diffrence Between Grand Total and 
            # Closing Stock Field into pivot sheet for tempalte
            pivotWorksheet = load_workbook(RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/60-Requirement-Summary/Requirement-Summary.xlsx")
            pivotSheet = pivotWorksheet.active

            pivotSheet.insert_rows(1,2)
            pivotSheet.insert_rows(6,2) # IGST/CGST Type
            
            
            df = pd.DataFrame(pivotSheet, index=None)
            rows = len(df.axes[0])
            cols = len(df.axes[1])
            pivotSheet.insert_rows(rows+1) # For Grand Total At bottom of the sheet

            

            for i in range(9,rows):
                pivotSheet.cell(i,cols-3).value = "=SUM(B"+str(i)+":"+openpyxl.utils.cell.get_column_letter(cols-4)+str(i)+")"
                pivotSheet.cell(i,cols-1).value = '='+openpyxl.utils.cell.get_column_letter(cols-2)+str(i)+'-'+openpyxl.utils.cell.get_column_letter(cols-3)+str(i)
                pivotSheet.cell(i,cols-2).value = "="+formulaSheet.cell(10,2).value.replace("#VAL#",str(i)) 

            
            pivotSheet.cell(6,1).value = 'IGST/CGST Type'
            pivotSheet.cell(7,1).value = 'Order No'
            pivotSheet.cell(rows+1,1).value = 'Grand Total'
            
            VAL = ''
            for j in range(2,cols-3):
                cellValue = "="+formulaSheet.cell(2,2).value.replace("#VAL#",openpyxl.utils.cell.get_column_letter(j))
                pivotSheet.cell(6,j).value = cellValue
                pivotSheet.cell(rows+1,j).value = "=SUM("+openpyxl.utils.cell.get_column_letter(j)+str(9)+":"+openpyxl.utils.cell.get_column_letter(j)+str(rows)+")"

            todayDate = datetime.today().strftime('%Y-%m-%d')
            pivotSheet.cell(2,3).value = 'Order Date'
            pivotSheet.cell(2,4).value = '-'


            pivotWorksheet.save(RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/60-Requirement-Summary/Requirement-Summary.xlsx")
            logger.info('Generated requirement summary file from consolidated orders for order date - '+OrderDate)
            formulaWorksheet.save(Formulasheet+'/FormulaSheet.xlsx')

            formulaWorksheet.close()
            return 'Generated Requirement Summary file'


    except Exception as e:
        logger.error("Error while generating Requirement-Summary file: "+str(e))
