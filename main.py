import log
import os
import shutil
from datetime import datetime
import pandas as pd
import xlsxwriter
import glob
import numpy as np
from openpyxl import load_workbook,Workbook
import openpyxl.utils.cell
import time
from config import ConfigFolderPath
import tabula
import csv
import json



logger = log.setup_custom_logger('root')


def downloadFiles(RootFolder,POSource,OrderDate,ClientCode):
    #converting str to datetime
    print(OrderDate)
    OrderDate = datetime.strptime(OrderDate, '%Y-%m-%d')
    # extracting year from the order date
    year = OrderDate.strftime("%Y")
    # formatting order date {2022-00-00) format
    OrderDate = OrderDate.strftime('%Y-%m-%d')
    source_folder = POSource
    destination_folder = RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/10-Download-Files/"
    try:
        for file_name in os.listdir(POSource):
            # construct full file path
            
            source = source_folder +"/"+ file_name
            destination = destination_folder
            # copy only files
            if os.path.isfile(source):

                shutil.copy(source, destination)
                logger.info("File '"+file_name+"' copied from source '"+source_folder+"' to destination '"+destination_folder+"'")
                print("File '"+file_name+"' copied from source '"+source_folder+"' to destination '"+destination_folder+"'")
    except Exception as e:
        logger.error("Error while copying files: "+str(e))
        print("Error while copying files: "+str(e))

def scriptStarted():
    logger.info('Starting script')
    print('Starting script')
    print('Starting is running...')
    print('Do not close this window while processing...')
    return "Script Started."

def scriptEnded():
    logger.info('Script Ended')
    print('Script Ended')
    return "Script Ended"
    
def checkFolderStructure(RootFolder,ClientCode,OrderDate):
    try:
        #converting str to datetime
        OrderDate = datetime.strptime(OrderDate, '%Y-%m-%d')

        # extracting year from the order date
        year = OrderDate.strftime("%Y")

        # formatting order date {2022-00-00) format
        OrderDate = OrderDate.strftime('%Y-%m-%d')
        
        # Checking if the folder exists or not, if doesnt exists, then script will create one.
        logger.info('Checking if the folder exists or not, if doesnt exists, then script will create one.')
        print('Checking if the folder exists or not, if doesnt exists, then script will create one.')
        DatedPath = RootFolder +"/"+ClientCode+"-"+year+"/"+str(OrderDate)
        isExist = os.path.exists(DatedPath)
        if not isExist:
            logger.info("Creating a new folder '"+ClientCode+"-"+year+"/"+str(OrderDate)+"' at location "+ RootFolder)
            print("Creating a new folder '"+ClientCode+"-"+year+"/"+str(OrderDate)+"' at location "+ RootFolder)
            os.makedirs(DatedPath)
            os.makedirs(DatedPath+"/10-Download-Files")
            os.makedirs(DatedPath+"/20-Intermediate-Files")
            os.makedirs(DatedPath+"/30-Extract-CSV")
            os.makedirs(DatedPath+"/40-Extract-Excel")
            os.makedirs(DatedPath+"/50-Consolidate-Orders")
            os.makedirs(DatedPath+"/60-Requirement-Summary")
            os.makedirs(DatedPath+"/70-Packaging-Slip")
            os.makedirs(DatedPath+"/80-Logs")
            logger.info("Folder '"+ClientCode+"-"+year+"/"+str(OrderDate)+"' is created.")
            print("Folder '"+ClientCode+"-"+year+"/"+str(OrderDate)+"' is created.")
        logger.info("Folder '"+ClientCode+"-"+year+"/"+str(OrderDate)+"' exists.")
        print("Folder '"+ClientCode+"-"+year+"/"+str(OrderDate)+"' exists.")

    except Exception as e:
        logger.error("Error while checking folder structure:  "+str(e))
        print("Error while checking folder structure:  "+str(e))

def mergeExcelsToOne(RootFolder,POSource,OrderDate,ClientCode): 
    #converting str to datetime
    OrderDate = datetime.strptime(OrderDate, '%Y-%m-%d')
    # extracting year from the order date
    year = OrderDate.strftime("%Y")
    # formatting order date {2022-00-00) format
    OrderDate = OrderDate.strftime('%Y-%m-%d')
    inputpath = RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/40-Extract-Excel/"
    outputpath = RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/50-Consolidate-Orders/"

    try:
        # logger.info('Checking 40-Extract-Excel directory exists or not.')
        file_list =  glob.glob(inputpath+ "/*.xlsx")
        if len(file_list) == 0:
            logger.info('No excel files found to merge.')
            print('No excel files found to merge.')
            return
        else:
            excl_list = []
            for f in os.listdir(inputpath):
                logger.info("Accessing '"+f+"' right now: ")
                print("Accessing '"+f+"' right now: ")
                df = pd.read_excel(inputpath+"/"+f)
                df.insert(0, "file_name", f)
                excl_list.append(df)
            excl_merged = pd.concat(excl_list, ignore_index=True,)
            excl_merged.to_excel(outputpath+"/"+'Consolidate-Orders.xlsx', index=False)
            logger.info("Merged "+str(len(file_list))+ " excel files to a single excel as 'Consolidate-Orders.xlsx'")
            print("Merged "+str(len(file_list))+ " excel files to a single excel as 'Consolidate-Orders.xlsx'")
            return 'All excels are merged into a single excel file'
    except Exception as e:
        logger.info("Error while merging files: "+str(e))
        print("Error while merging files: "+str(e))


def mergeToPivot(RootFolder,POSource,OrderDate,ClientCode,Formulasheet):


    with open(ConfigFolderPath+'client.json', 'r') as jsonFile:
        config = json.load(jsonFile)
        ClientName = config

    try:
        #converting str to datetime
        OrderDate = datetime.strptime(OrderDate, '%Y-%m-%d')
        # extracting year from the order date
        year = OrderDate.strftime("%Y")
        # formatting order date {2022-00-00) format
        OrderDate = OrderDate.strftime('%Y-%m-%d')

        formulaWorksheet = load_workbook(Formulasheet+'/FormulaSheet.xlsx',data_only=True) 
        # Data_only = True is used to get evaluated formula value instead of formula
        formulaSheet = formulaWorksheet['FormulaSheet']
        if not os.path.exists(RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/50-Consolidate-Orders/Consolidate-Orders.xlsx"):
            logger.info("Could not find the consolidate order folder to generate requirement summary file")
            print("Could not find the consolidate order folder to generate requirement summary file")
            return
        else:
            df = pd.read_excel(RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/50-Consolidate-Orders/Consolidate-Orders.xlsx")
            df_pivot = pd.pivot_table(df, index="ArticleEAN", values='Qty', 
            columns=['Vendor Name','PO Number','Receiving Location'], aggfunc='sum')
            df_pivot['Grand Total'] = 0
            df_pivot['Closing Stock'] = 0
            df_pivot['Diff CS - GT'] = 0
            df_pivot['Rate'] = 0
            df_pivot.to_excel(RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/60-Requirement-Summary/Requirement-Summary.xlsx")


            # Adding Processing Date, Order Number and Closing Stock, Diffrence Between Grand Total and 
            # Closing Stock Field into pivot sheet for tempalte
            pivotWorksheet = load_workbook(RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/60-Requirement-Summary/Requirement-Summary.xlsx")
            pivotSheet = pivotWorksheet.active

            pivotSheet.insert_rows(1,2)
            pivotSheet.insert_rows(6,2) # IGST/CGST Type
            
            
            df = pd.DataFrame(pivotSheet, index=None)
            rows = len(df.axes[0])
            cols = len(df.axes[1])
            pivotSheet.insert_rows(rows+1) # For Grand Total At bottom of the sheet

            

            for i in range(9,rows):
                pivotSheet.cell(i,cols-3).value = "=SUM(B"+str(i)+":"+openpyxl.utils.cell.get_column_letter(cols-4)+str(i)+")"
                pivotSheet.cell(i,cols-1).value = '='+openpyxl.utils.cell.get_column_letter(cols-2)+str(i)+'-'+openpyxl.utils.cell.get_column_letter(cols-3)+str(i)
                pivotSheet.cell(i,cols-2).value = "="+formulaSheet.cell(10,2).value.replace("#VAL#",str(i)) 

            
            pivotSheet.cell(6,1).value = 'IGST/CGST Type'
            pivotSheet.cell(7,1).value = 'Order No'
            pivotSheet.cell(rows+1,1).value = 'Grand Total'
            
            VAL = ''
            for j in range(2,cols-3):
                cellValue = "="+formulaSheet.cell(2,2).value.replace("#VAL#",openpyxl.utils.cell.get_column_letter(j))
                pivotSheet.cell(6,j).value = cellValue
                pivotSheet.cell(rows+1,j).value = "=SUM("+openpyxl.utils.cell.get_column_letter(j)+str(9)+":"+openpyxl.utils.cell.get_column_letter(j)+str(rows)+")"

            todayDate = datetime.today().strftime('%Y-%m-%d')
            pivotSheet.cell(2,3).value = 'Order Date'
            pivotSheet.cell(2,4).value = OrderDate

            pivotSheet.cell(2,5).value = 'ClientName'
            pivotSheet.cell(2,6).value = ClientName[ClientCode]



            pivotWorksheet.save(RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/60-Requirement-Summary/Requirement-Summary.xlsx")
            logger.info('Generated requirement summary file from consolidated orders for order date - '+OrderDate)
            print('Generated requirement summary file from consolidated orders for order date - '+OrderDate)
            formulaWorksheet.save(Formulasheet+'/FormulaSheet.xlsx')

            formulaWorksheet.close()
            return 'Generated Requirement Summary file'


    except Exception as e:
        logger.error("Error while generating Requirement-Summary file: "+str(e))
        print("Error while generating Requirement-Summary file: "+str(e))


def getFilesToProcess(RootFolder,POSource,OrderDate,ClientCode):
    #converting str to datetime
    OrderDate = datetime.strptime(OrderDate, '%Y-%m-%d')
    # extracting year from the order date
    year = OrderDate.strftime("%Y")
    # formatting order date {2022-00-00) format
    OrderDate = OrderDate.strftime('%Y-%m-%d')
    try:
        inputFolderPath = RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/10-Download-Files/"
        outputFolderPath = RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/40-Extract-Excel/"
        startedProcessing = time.time()
        
        if len(os.listdir(inputFolderPath)) == 0:
            logger.info("'"+inputFolderPath+"' Folder is empty, add pdf files to convert")
            print("'"+inputFolderPath+"' Folder is empty, add pdf files to convert")
            return
        else:
            for f in os.listdir(inputFolderPath):
                fOutputExtension = f.replace('.pdf', '.xlsx')
                
                pdfToTable(inputFolderPath+f,outputFolderPath+fOutputExtension,RootFolder,POSource,OrderDate,ClientCode,f)
            
            print("Successfully Converted all the PDF Files to Excel Files!")
            print("Conversion Completed in "+"{:.2f}".format(time.time() - startedProcessing,2)+ " seconds!")
    except Exception as e:
        logger.error("Error while processing files: "+str(e))
        print("Error while processing files: "+str(e))


def pdfToTable(inputPath,outputPath,RootFolder,POSource,OrderDate,ClientCode,filecsv):

    try:
        logger.info("Converting '"+ filecsv +"' to excel '"+filecsv.replace('.pdf', '.xlsx')+"'")
        print("Converting '"+ filecsv +"' to excel '"+filecsv.replace('.pdf', '.xlsx')+"'")
        #converting str to datetime
        OrderDate = datetime.strptime(OrderDate, '%Y-%m-%d')
        # extracting year from the order date
        year = OrderDate.strftime("%Y")
        # formatting order date {2022-00-00) format
        OrderDate = OrderDate.strftime('%Y-%m-%d')

        startedProcessing = time.time()

        intermediateCSV = RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/30-Extract-CSV/"+filecsv.replace('.pdf', '.csv')
        intermediateExcel = RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/20-Intermediate-Files/1_"+filecsv.replace('.pdf', '.xlsx')
        intermediateExcel2 = RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/20-Intermediate-Files/2_"+filecsv.replace('.pdf', '.xlsx')
        intermediateoutputPath = RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/20-Intermediate-Files/3_"+filecsv.replace('.pdf', '.xlsx')

        tabula.convert_into(input_path= inputPath , output_path= intermediateCSV , pages = 'all', lattice= True)

        # making a dataframe using csv
        df = pd.read_csv(filepath_or_buffer= intermediateCSV, skiprows =[1,2,3])


        # converting csv to excel
        writer = pd.ExcelWriter(intermediateExcel, engine='xlsxwriter')
        df.to_excel(writer, sheet_name= 'Sheet1', merge_cells= False)
        writer.save()


        # loading excel
        workbook = load_workbook(filename= intermediateExcel)
        sheet = workbook.active


        # finding the last row needed
        endrow = 1
        endcolumn = 1

        for i in range(1, 10000000):
            if sheet.cell(i,1).value != None and "Grand" in sheet.cell(i,1).value:
                endrow = i
                break

        for i in range(1, 10000000):
            if sheet.cell(2,i).value != None and "Total" in sheet.cell(2,i).value:
                endcolumn = i
                break

        # print(endrow, endcolumn)
            


        # Cleaning loop
        for i in range(1, endrow+2):
            for j in range(1, endcolumn+1):
                if sheet.cell(i,j).value != None and "Aditya" in sheet.cell(i,j).value:
                    sheet.cell(i,1).value = "remove"
                elif sheet.cell(i,j).value != None and "P. O. Number" in sheet.cell(i,j).value:
                    sheet.cell(i,1).value = "remove"
                elif sheet.cell(i,j).value != None and "PO" in sheet.cell(i,j).value and i != 2:
                        sheet.cell(i,1).value = "remove"
                elif sheet.cell(i,j).value != None and "_x000D_" in sheet.cell(i,j).value:
                    temp = sheet.cell(i,j).value.replace("_x000D_", "")
                    sheet.cell(i,j).value = temp


        # rearranging the data into a single row

        # column numbers
        CGSTRate = endcolumn+1
        CGSTAmt = endcolumn+2
        SGSTRate = endcolumn+3
        SGSTAmt = endcolumn+4
        UTGSTRate = endcolumn+5
        UTGSTAmt = endcolumn+6
        IGSTRate = endcolumn+7
        IGSTAmt = endcolumn+8
        VendorPenalty = endcolumn+9
        VendorName = endcolumn+8
        ReceivingLocation = endcolumn+9
        PONumber = endcolumn+10
        VendorCode = endcolumn+11
        for i in range(1, endrow):
            for j in range(1, endcolumn+1):
                if sheet.cell(i,j).value == "CGST":
                    sheet.cell(i-1,CGSTRate).value = sheet.cell(i,j+1).value
                    sheet.cell(i-1,CGSTAmt).value = sheet.cell(i,j+2).value
                    sheet.cell(i,j+1).value = ""
                    sheet.cell(i,j+2).value = ""
                    sheet.cell(i,j).value = ""
                elif sheet.cell(i,j).value == "SGST":
                    sheet.cell(i-2,SGSTRate).value = sheet.cell(i,j+1).value
                    sheet.cell(i-2,SGSTAmt).value = sheet.cell(i,j+2).value
                    sheet.cell(i,j+1).value = ""
                    sheet.cell(i,j+2).value = ""
                    sheet.cell(i,j).value = ""
                elif sheet.cell(i,j).value == "UTGST":
                    sheet.cell(i-3,UTGSTRate).value = sheet.cell(i,j+1).value
                    sheet.cell(i-3,UTGSTAmt).value = sheet.cell(i,j+2).value
                    sheet.cell(i,j+1).value = ""
                    sheet.cell(i,j+2).value = ""
                    sheet.cell(i,j).value = ""
                elif sheet.cell(i,j).value == "IGST":
                    sheet.cell(i-4,IGSTRate).value = sheet.cell(i,j+1).value
                    sheet.cell(i-4,IGSTAmt).value = sheet.cell(i,j+2).value
                    sheet.cell(i,j+1).value = ""
                    sheet.cell(i,j+2).value = ""
                    sheet.cell(i,j).value = ""
                elif sheet.cell(i,j).value == "VendorPenalty":
                    sheet.cell(i-4,VendorPenalty).value = sheet.cell(i,j+1).value
                    sheet.cell(i,j+1).value = ""
                    sheet.cell(i,j).value = ""


        # removing empty and unwanted rows
        counter = endrow
        while counter > 0:
            if sheet.cell(counter,1).value == None or sheet.cell(counter,1).value == "" or sheet.cell(counter, 1).value == "remove":
                sheet.delete_rows(counter)
            counter = counter -1

        sheet.delete_cols(endcolumn-1)
        sheet.delete_cols(endcolumn-2)

        sheet["A1"] = "POItem"
        sheet["B1"] = "ArticleEAN"
        sheet["C1"] = "Article Number"
        sheet["D1"] = "ArticleDescription"
        sheet["E1"] = "HSNCode"
        sheet["F1"] = "MRP"
        sheet["G1"] = "BasicCostPrice(TaxableValue)"
        sheet["H1"] = "Qty"
        sheet["I1"] = "UM"
        sheet["J1"] = "TaxableValue"
        # sheet["K1"] = "GSTRate"
        # sheet["L1"] = "GSTAmt"
        sheet["K1"] = "Total Amount"
        sheet["L1"] = "CGSTRate"
        sheet["M1"] = "CGSTAmt"
        sheet["N1"] = "SGSTRate"
        sheet["O1"] = "SGSTAmt"
        sheet["P1"] = "UTGSTRate"
        sheet["Q1"] = "UTGSTAmt"
        sheet["R1"] = "IGSTRate"
        sheet["S1"] = "IGSTAmt"
        sheet["T1"] = "Vendor Penalty"
        sheet["U1"] = "Vendor Name"
        sheet["V1"] = "Receiving Location"
        sheet["W1"] = "PO Number"
        sheet["X1"] = "Vendor Code"

        # workbook = xlsxwriter.Workbook(intermediateExcel, {'strings_to_numbers': True})
        df2 = pd.read_csv(filepath_or_buffer= intermediateCSV, on_bad_lines='skip')

        # print(df2)
        writer2 = pd.ExcelWriter(intermediateExcel2, engine='xlsxwriter')
        df2.to_excel(writer2, sheet_name= 'Sheet1', merge_cells= False)
        writer2.save()
        workbook2 = load_workbook(filename= intermediateExcel2)
        sheet2 = workbook2.active

        # Cleaning loop
        for i in range(1, 10):
            for j in range(1, 5):
                if sheet2.cell(i,j).value != None and "_x000D_" in sheet2.cell(i,j).value:
                    temp = sheet2.cell(i,j).value.replace("_x000D_", "%")
                    sheet2.cell(i,j).value = temp

        # Recalculating endrow
        for i in range(1, 10000000):
            if sheet.cell(i,1).value != None and "Grand" in sheet.cell(i,1).value:
                endrow = i
                break

        PONumberRow = 1
        for i in range(1, 25):
            if sheet2.cell(i,1).value != None and "P. O. Number" in sheet2.cell(i,1).value:
                PONumberRow = i
                break

        VendorNameValue = sheet2.cell(2, 1).value.replace(":","").split('%')[1]
        VendorCodeValue = sheet2.cell(2, 1).value.replace(":","").split('%')[3]
        ReceivingLocationValue = sheet2.cell(4, 1).value.replace(":","").split('%')[3]
        PONumValue = sheet2.cell(PONumberRow, 1).value.replace(":", "").split('%')[1]
        for i in range(2, endrow):
            sheet.cell(i, VendorName).value = VendorNameValue
            sheet.cell(i, ReceivingLocation).value = ReceivingLocationValue
            sheet.cell(i, PONumber).value = PONumValue
            sheet.cell(i, VendorCode).value = VendorCodeValue
        
        # get  Delete row count
        lastrow = 1
        for i in range(endrow, 1000000):
            if sheet.cell(i,1).value != None and "Other Conditions" in sheet.cell(i,1).value:
                lastrow = i+5
                break
            
        while endrow <= lastrow:
            sheet.delete_rows(lastrow)
            lastrow = lastrow -1

        workbook2.save(filename= intermediateoutputPath)
        workbook.save(filename= outputPath)

        logger.info("Converted '"+ inputPath + "' to '" + outputPath+"'"+ " in "+ "{:.2f}".format(time.time() - startedProcessing,2)+ " seconds.")
        print("Converted '"+ inputPath + "' to '" + outputPath+"'"+ " in "+ "{:.2f}".format(time.time() - startedProcessing,2)+ " seconds.")
        return "Conversion Complete!"
    
    except Exception as e:
        logger.error("Error while processing file: "+str(e))
        print("Error while processing file: "+str(e))


def generatingPackaingSlip(RootFolder,ReqSource,OrderDate,ClientCode,Formulasheet,TemplateFiles):
    try:
        #converting str to datetime
        OrderDate = datetime.strptime(OrderDate, '%Y-%m-%d')
        # extracting year from the order date
        year = OrderDate.strftime("%Y")
        # formatting order date {2022-00-00) format
        OrderDate = OrderDate.strftime('%Y-%m-%d')

        startedTemplating = time.time()
        sourcePivot = ReqSource
        source = TemplateFiles+"/PackingSlip-Template.xlsx"
        destination = TemplateFiles+"/TemplateFile.xlsx"
        

        # Making Copy of template file
        # shutil.copy(source, destination)
        # print("File copied successfully.")

        # Load work vook and sheets
        InputWorkbook = load_workbook(sourcePivot,data_only=True)
        # TemplateWorkbook = load_workbook(destination)

        InputSheet = InputWorkbook.active
        # TemplateSheet = TemplateWorkbook['ORDER']

        # Get rows and Column count
        df = pd.DataFrame(InputSheet, index=None)
        rows = len(df.axes[0])
        cols = len(df.axes[1])


        formulaWorksheet = load_workbook(Formulasheet+'/FormulaSheet.xlsx',data_only=True)
        formulaSheet = formulaWorksheet['FormulaSheet']
        DBFformula = formulaWorksheet['DBF']
        
        
        
        for column in range(2,cols-3):
            startedTemplatingFile = time.time()
            # Making Copy of template file
            shutil.copy(source, destination)
            logger.info("Template File copied successfully for generating packaging-slip")
            print("Template File copied successfully for generating packaging-slip")

            # Load work vook and sheets
            TemplateWorkbook = load_workbook(destination, data_only=True)
            TemplateSheet = TemplateWorkbook['ORDER']
            dbfsheet = TemplateWorkbook['DBF']

            
            TemplateSheet.cell(6,4).value = InputSheet.cell(7,2).value # Order Name
            TemplateSheet.cell(5,1).value = InputSheet.cell(7,2).value # Order Name
            # PO Number
            filename = InputSheet.cell(4,column).value
            TemplateSheet.cell(5,2).value = InputSheet.cell(4,column).value
            TemplateSheet.cell(6,2).value = InputSheet.cell(4,column).value
            TemplateSheet.cell(6,1).value = InputSheet.cell(4,column).value
            TemplateSheet.cell(6,3).value = InputSheet.cell(4,column).value
            # Receving Location
            TemplateSheet.cell(5,3).value = InputSheet.cell(5,column).value
            TemplateSheet.cell(4,2).value = InputSheet.cell(5,column).value

            TemplateSheet.cell(1,1).value = 'DATE'
            TemplateSheet.cell(1,2).value = InputSheet.cell(2,4).value # Date

            TemplateSheet.cell(1,3).value = 'SGST/IGST'
            TemplateSheet.cell(1,4).value = InputSheet.cell(6,column).value # IGST/SGST Type
            print(TemplateSheet.cell(1,4).value,InputSheet.cell(6,column).value)

            # Copy EAN to template sheet
            Trows = 8
            Tcols = 5
            dbfrows = 2
            dbfcols = 57
            for row in range(7,rows):
                # if InputSheet.cell(row,column).value != None or InputSheet.cell(row,column).value != "":
                if str(InputSheet.cell(row,column).value).isnumeric():
                    
                        
                    # Copy Qty to template sheet
                    TemplateSheet.cell(Trows,Tcols).value = InputSheet.cell(row,column).value 
                    TemplateSheet.cell(Trows,Tcols+1).value = InputSheet.cell(row,column).value
                    TemplateSheet.cell(Trows,Tcols+2).value = "="+openpyxl.utils.cell.get_column_letter(Tcols+1)+str(Trows) # Actual Qty
                    
                    # Copy EAN to template sheet
                    TemplateSheet.cell(Trows,Tcols-3).value = InputSheet.cell(row,1).value

                    # VLOOKUP
                    # StyleName
                    TemplateSheet.cell(Trows,Tcols-4).value = "="+formulaSheet.cell(3,2).value.replace("#VAL#",str(Trows))

                    # style
                    TemplateSheet.cell(Trows,Tcols-2).value =  "="+formulaSheet.cell(4,2).value.replace("#VAL#",str(Trows))

                    # SADM SKU
                    TemplateSheet.cell(Trows,Tcols-1).value = "="+formulaSheet.cell(5,2).value.replace("#VAL#",str(Trows))
                    
                    # Rate (in Rs.) Order file
                    TemplateSheet.cell(Trows,Tcols+3).value = "="+formulaSheet.cell(6,2).value.replace("#VAL#",str(Trows))

                    #Cls stk vs order
                    # TemplateSheet.cell(Trows,Tcols+6).value = TemplateSheet.cell(Trows,Tcols+5).value - TemplateSheet.cell(Trows,Tcols+2).value
                    TemplateSheet.cell(Trows,Tcols+6).value = "="+openpyxl.utils.cell.get_column_letter(Tcols+5)+str(Trows) +'-'+openpyxl.utils.cell.get_column_letter(Tcols+2)+str(Trows) 

                    # LOCATION2
                    TemplateSheet.cell(Trows,Tcols+7).value = "="+formulaSheet.cell(7,2).value.replace("#VAL#",str(Trows))

                    #BULK  / DTA  BULK  /  EOSS LOC
                    TemplateSheet.cell(Trows,Tcols+8).value = "="+formulaSheet.cell(8,2).value.replace("#VAL#",str(Trows))

                    #MRP
                    TemplateSheet.cell(Trows,Tcols+9).value = "="+formulaSheet.cell(9,2).value.replace("#VAL#",str(Trows))

                    # Closing stk
                    TemplateSheet.cell(Trows,Tcols+5).value = "="+formulaSheet.cell(11,2).value.replace("#VAL#",str(Trows)) 

                    #SCAN
                    TemplateSheet.cell(Trows,Tcols+10).value = "="+formulaSheet.cell(12,2).value.replace("#VAL#",str(Trows))

                    #SCAN VS DIFF
                    TemplateSheet.cell(Trows,Tcols+11).value = "="+formulaSheet.cell(13,2).value.replace("#VAL#",str(Trows))

                    # ERROR
                    TemplateSheet.cell(Trows,Tcols+12).value = "="+formulaSheet.cell(14,2).value.replace("#VAL#",str(Trows))

                    # STYLE COLOR
                    TemplateSheet.cell(Trows,Tcols+13).value = "="+formulaSheet.cell(15,2).value.replace("#VAL#",str(Trows))
                    
                    # Adding values to DBF
                    # for i in range(1,dbfcols):
                    #     dbfsheet.cell(dbfrows, i).value = '='+DBFformula.cell(2,i+1).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows))
                    
                    dbfsheet.cell(dbfrows, 1).value = '='+DBFformula.cell(2,2).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #Vouchertypename
                    dbfsheet.cell(dbfrows, 2).value = '='+DBFformula.cell(2,3).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #CSNNO
                    dbfsheet.cell(dbfrows, 3).value = '='+DBFformula.cell(2,4).value.replace("#VAL#",str(Trows)) #Date
                    dbfsheet.cell(dbfrows, 4).value = '='+DBFformula.cell(2,5).value.replace("#VAL#",str(Trows)) #Reference
                    dbfsheet.cell(dbfrows, 5).value = '='+DBFformula.cell(2,6).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #REF1  
                    dbfsheet.cell(dbfrows, 6).value = '='+DBFformula.cell(2,7).value.replace("#VAL#",str(Trows)) #Dealer Name
                    dbfsheet.cell(dbfrows, 7).value = '='+DBFformula.cell(2,8).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) # PriceLevel
                    dbfsheet.cell(dbfrows, 8).value = '='+DBFformula.cell(2,9).value.replace("#VAL#",str(Trows)) #ItemName/ SKU
                    dbfsheet.cell(dbfrows, 9).value = '='+DBFformula.cell(2,10).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #GODOWN
                    dbfsheet.cell(dbfrows, 10).value = '='+DBFformula.cell(2,11).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #Qty
                    dbfsheet.cell(dbfrows, 11).value = '='+DBFformula.cell(2,12).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #Rate
                    dbfsheet.cell(dbfrows, 12).value = '='+DBFformula.cell(2,13).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #SUBTOTAL
                    dbfsheet.cell(dbfrows, 13).value = '='+DBFformula.cell(2,14).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #DISCPERC
                    dbfsheet.cell(dbfrows, 14).value = '='+DBFformula.cell(2,15).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #DISCAMT
                    dbfsheet.cell(dbfrows, 15).value = '='+DBFformula.cell(2,16).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #ITEMVALUE
                    dbfsheet.cell(dbfrows, 16).value = '='+DBFformula.cell(2,17).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #LedgerAcct
                    dbfsheet.cell(dbfrows, 17).value = '='+DBFformula.cell(2,18).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #CATEGORY1
                    dbfsheet.cell(dbfrows, 18).value = '='+DBFformula.cell(2,19).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #COSTCENT1
                    dbfsheet.cell(dbfrows, 19).value = '='+DBFformula.cell(2,20).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #CATEGORY2
                    dbfsheet.cell(dbfrows, 20).value = '='+DBFformula.cell(2,21).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #COSTCENT2
                    dbfsheet.cell(dbfrows, 21).value = '='+DBFformula.cell(2,22).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #CATEGORY3
                    dbfsheet.cell(dbfrows, 22).value = '='+DBFformula.cell(2,23).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #COSTCENT3
                    dbfsheet.cell(dbfrows, 23).value = '='+DBFformula.cell(2,24).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #CATEGORY4
                    dbfsheet.cell(dbfrows, 24).value = '='+DBFformula.cell(2,25).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #COSTCENT4
                    dbfsheet.cell(dbfrows, 25).value = '='+DBFformula.cell(2,26).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #ITEMTOTAL
                    dbfsheet.cell(dbfrows, 26).value = '='+DBFformula.cell(2,27).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #TOTALQTY
                    dbfsheet.cell(dbfrows, 27).value = '='+DBFformula.cell(2,28).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #CDISCHEAD
                    dbfsheet.cell(dbfrows, 28).value = '='+DBFformula.cell(2,29).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #CDISCPERC
                    dbfsheet.cell(dbfrows, 29).value = '='+DBFformula.cell(2,30).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #COMMONDISC
                    dbfsheet.cell(dbfrows, 30).value = '='+DBFformula.cell(2,31).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #BEFORETAX
                    dbfsheet.cell(dbfrows, 31).value = '='+DBFformula.cell(2,32).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #TAXHEAD
                    dbfsheet.cell(dbfrows, 32).value = '='+DBFformula.cell(2,33).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #TAXPERC
                    dbfsheet.cell(dbfrows, 33).value = '='+DBFformula.cell(2,34).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #TAXAMT
                    dbfsheet.cell(dbfrows, 34).value = '='+DBFformula.cell(2,35).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #STAXHEAD
                    dbfsheet.cell(dbfrows, 35).value = '='+DBFformula.cell(2,36).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #STAXPERC
                    dbfsheet.cell(dbfrows, 36).value = '='+DBFformula.cell(2,37).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #STAXAMT
                    dbfsheet.cell(dbfrows, 37).value = '='+DBFformula.cell(2,38).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #ITAXHEAD
                    dbfsheet.cell(dbfrows, 38).value = '='+DBFformula.cell(2,39).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #ITAXPERC
                    dbfsheet.cell(dbfrows, 39).value = '='+DBFformula.cell(2,40).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #ITAXAMT
                    dbfsheet.cell(dbfrows, 40).value = '='+DBFformula.cell(2,41).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #NETAMT
                    dbfsheet.cell(dbfrows, 41).value = '='+DBFformula.cell(2,42).value.replace("#VAL#",str(Trows)) #ROUND
                    dbfsheet.cell(dbfrows, 42).value = '='+DBFformula.cell(2,43).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #ROUND1
                    dbfsheet.cell(dbfrows, 43).value = '='+DBFformula.cell(2,44).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #REFTYPE
                    dbfsheet.cell(dbfrows, 44).value = '='+DBFformula.cell(2,45).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #Name
                    dbfsheet.cell(dbfrows, 45).value = '='+DBFformula.cell(2,46).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #REFAMT
                    dbfsheet.cell(dbfrows, 46).value = '='+DBFformula.cell(2,47).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #Narration
                    dbfsheet.cell(dbfrows, 47).value = '='+DBFformula.cell(2,48).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #Transport
                    dbfsheet.cell(dbfrows, 48).value = '='+DBFformula.cell(2,49).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #transmode
                    dbfsheet.cell(dbfrows, 49).value = '='+DBFformula.cell(2,50).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #pymtterm
                    dbfsheet.cell(dbfrows, 50).value = '='+DBFformula.cell(2,51).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #OrderNo/PO number
                    dbfsheet.cell(dbfrows, 51).value = '='+DBFformula.cell(2,52).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #orddate
                    dbfsheet.cell(dbfrows, 52).value = '='+DBFformula.cell(2,53).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #DANO
                    dbfsheet.cell(dbfrows, 53).value = '='+DBFformula.cell(2,54).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #Delyadd1
                    dbfsheet.cell(dbfrows, 54).value = '='+DBFformula.cell(2,55).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #Delyadd2
                    dbfsheet.cell(dbfrows, 55).value = '='+DBFformula.cell(2,56).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #Delyadd3
                    dbfsheet.cell(dbfrows, 56).value = '='+DBFformula.cell(2,57).value.replace("#VAL#",str(Trows)).replace("#DBFROWS#",str(dbfrows)) #Delyadd4



                    Trows += 1
                    dbfrows += 1 

            
            TemplateWorkbook.save(RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/70-Packaging-Slip/"+"PackagingSlip_"+str(filename)+".xlsx")
            logger.info("Packaing slip generated for: "+str(filename)+ " file - {:.2f}".format(time.time() - startedTemplatingFile,2)+ " seconds.")
            print("Packaing slip generated for: "+str(filename)+ " file - {:.2f}".format(time.time() - startedTemplatingFile,2)+ " seconds.")
            
        logger.info("Total time taken for generation of packaging-slips:  {:.2f}".format(time.time() - startedTemplating,2)+ " seconds.")
        print("Total time taken for generation of packaging-slips:  {:.2f}".format(time.time() - startedTemplating,2)+ " seconds.")
        return 'Completed!'
    except Exception as e:
        logger.error("Error while generating packaging-slip file: "+str(e))
        print("Error while generating packaging-slip file: "+str(e))